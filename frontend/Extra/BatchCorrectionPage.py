import json
import os
import re
import shutil
import threading

import openpyxl
import openpyxl.worksheet.worksheet
from PyQt5.QtCore import Qt
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtWidgets import QLayout
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QWidget
from qfluentwidgets import FluentIcon
from qfluentwidgets import FluentWindow
from qfluentwidgets import PushButton
from qfluentwidgets import SingleDirectionScrollArea
from qfluentwidgets import TransparentPushButton

from base.Base import Base
from model.Item import Item
from module.Config import Config
from module.File.FileManager import FileManager
from module.Localizer.Localizer import Localizer
from module.Localizer.LocalizerEN import LocalizerEN
from module.Localizer.LocalizerZH import LocalizerZH
from module.TableManager import TableManager
from widget.CommandBarCard import CommandBarCard
from widget.EmptyCard import EmptyCard

class BatchCorrectionPage(QWidget, Base):

    SINGLE: tuple[Item.FileType] = (
        Item.FileType.MD,
        Item.FileType.TXT,
        Item.FileType.ASS,
        Item.FileType.SRT,
        Item.FileType.EPUB,
        Item.FileType.MESSAGEJSON,
    )

    DOUBLE: tuple[Item.FileType] = (
        Item.FileType.XLSX,
        Item.FileType.WOLFXLSX,
        Item.FileType.RENPY,
        Item.FileType.TRANS,
        Item.FileType.KVJSON,
    )

    FILE_NAME_WHITELIST: re.Pattern = re.compile(r"^(结果检查_|result_check_)([^\\/]+)\.json$", flags = re.IGNORECASE)
    FILE_NAME_BLACKLIST: tuple[str] = (
        LocalizerEN.path_result_check_untranslated,
        LocalizerZH.path_result_check_untranslated,
    )

    def __init__(self, text: str, window: FluentWindow) -> None:
        super().__init__(window)
        self.setObjectName(text.replace(" ", "-"))

        # 载入并保存默认配置
        config = Config().load().save()

        # 设置主容器
        self.root = QVBoxLayout(self)
        self.root.setSpacing(8)
        self.root.setContentsMargins(24, 24, 24, 24) # 左、上、右、下

        # 添加控件
        self.add_widget_head(self.root, config, window)
        self.add_widget_body(self.root, config, window)
        self.add_widget_foot(self.root, config, window)

    # 头部
    def add_widget_head(self, parent: QLayout, config: Config, window: FluentWindow) -> None:
        parent.addWidget(
            EmptyCard(
                title = Localizer.get().batch_correction_page,
                description = Localizer.get().batch_correction_page_desc,
                init = None,
            )
        )

    # 主体
    def add_widget_body(self, parent: QLayout, config: Config, window: FluentWindow) -> None:
        # 创建滚动区域的内容容器
        scroll_area_vbox_widget = QWidget()
        scroll_area_vbox = QVBoxLayout(scroll_area_vbox_widget)
        scroll_area_vbox.setContentsMargins(0, 0, 0, 0)

        # 创建滚动区域
        scroll_area = SingleDirectionScrollArea(orient = Qt.Orientation.Vertical)
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(scroll_area_vbox_widget)
        scroll_area.enableTransparentBackground()

        # 将滚动区域添加到父布局
        parent.addWidget(scroll_area)

        # 添加控件
        self.add_step_01(scroll_area_vbox, config, window)
        self.add_step_02(scroll_area_vbox, config, window)
        scroll_area_vbox.addStretch(1)

    # 底部
    def add_widget_foot(self, parent: QLayout, config: Config, window: FluentWindow) -> None:
        self.command_bar_card = CommandBarCard()
        parent.addWidget(self.command_bar_card)

        # 添加命令
        self.command_bar_card.add_stretch(1)
        self.add_command_bar_action_wiki(self.command_bar_card, config, window)

    # WiKi
    def add_command_bar_action_wiki(self, parent: CommandBarCard, config: Config, window: FluentWindow) -> None:
        push_button = TransparentPushButton(FluentIcon.HELP, Localizer.get().wiki)
        push_button.clicked.connect(lambda: QDesktopServices.openUrl(QUrl("https://github.com/neavo/LinguaGacha/wiki")))
        parent.add_widget(push_button)

    # 第一步
    def add_step_01(self, parent: QLayout, config: Config, window: FluentWindow) -> None:

        def init(widget: EmptyCard) -> None:
            push_button = PushButton(FluentIcon.PLAY, Localizer.get().start)
            push_button.clicked.connect(lambda: self.step_01_clicked(window))
            widget.add_widget(push_button)

        widget = EmptyCard(
            title = Localizer.get().batch_correction_page_step_01,
            description = Localizer.get().batch_correction_page_step_01_desc,
            init = init,
        )
        parent.addWidget(widget)

    # 第二步
    def add_step_02(self, parent: QLayout, config: Config, window: FluentWindow) -> None:

        def init(widget: EmptyCard) -> None:
            push_button = PushButton(FluentIcon.SAVE_AS, Localizer.get().inject)
            push_button.clicked.connect(lambda: self.step_02_clicked(window))
            widget.add_widget(push_button)

        parent.addWidget(EmptyCard(
            title = Localizer.get().batch_correction_page_step_02,
            description = Localizer.get().batch_correction_page_step_02_desc,
            init = init,
        ))

    # WiKi
    def add_command_bar_action_wiki(self, parent: CommandBarCard, config: Config, window: FluentWindow) -> None:
        push_button = TransparentPushButton(FluentIcon.HELP, Localizer.get().wiki)
        push_button.clicked.connect(lambda: QDesktopServices.openUrl(QUrl("https://github.com/neavo/LinguaGacha/wiki")))
        parent.add_widget(push_button)

    # 第一步点击事件
    def step_01_clicked(self, window: FluentWindow) -> None:
        config = Config().load()

        data_dict: dict[str, dict] = {}
        for entry in os.scandir(config.input_folder):
            if (
                entry.is_file()
                and BatchCorrectionPage.FILE_NAME_WHITELIST.search(entry.name) is not None
                and entry.name not in BatchCorrectionPage.FILE_NAME_BLACKLIST
            ):
                with open(entry.path, "r", encoding = "utf-8-sig") as reader:
                    json_data: dict[str, dict[str, str]] = json.load(reader)

                    # 有效性校验
                    if not isinstance(json_data, dict):
                        continue

                    for file_path, items_by_path in json_data.items():
                        if not isinstance(items_by_path, dict):
                            continue

                        # 分别处理两种文本组织形式
                        chunks: list[str] = file_path.split("|")
                        if len(chunks) == 1:
                            group = BatchCorrectionPage.FILE_NAME_WHITELIST.sub(r"\2", entry.name)
                        else:
                            group = BatchCorrectionPage.FILE_NAME_WHITELIST.sub(r"\2", entry.name) + " " + chunks[1].strip()
                            file_path = chunks[0].strip()

                        # 添加数据
                        for src, dst in items_by_path.items():
                            key = (file_path, src)
                            data_dict.setdefault(key, {})["src"] = src
                            data_dict.setdefault(key, {})["dst"] = dst
                            data_dict.setdefault(key, {})["group"] = data_dict.get(key).get("group", []) + [group]
                            data_dict.setdefault(key, {})["file_path"] = file_path

        # 有效性检查
        if len(data_dict) == 0:
            self.emit(Base.Event.TOAST, {
                "type": Base.ToastType.ERROR,
                "message": Localizer.get().alert_no_data,
            })
            return None

        # 排序
        items: list[dict[str, str | list[str]]] = sorted(
            data_dict.values(),
            key = lambda x: (x.get("path"), x.get("group"))
        )

        # 新建工作表
        book: openpyxl.Workbook = openpyxl.Workbook()
        sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

        # 添加表头
        TableManager.set_cell_value(sheet, 1, 1, Localizer.get().batch_correction_page_title_01)
        TableManager.set_cell_value(sheet, 1, 2, Localizer.get().batch_correction_page_title_02)
        TableManager.set_cell_value(sheet, 1, 3, Localizer.get().batch_correction_page_title_03)
        TableManager.set_cell_value(sheet, 1, 4, Localizer.get().batch_correction_page_title_04)
        TableManager.set_cell_value(sheet, 1, 5, Localizer.get().batch_correction_page_title_05)

        # 设置表头
        sheet.auto_filter.ref = "A1:E1"
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 64
        sheet.column_dimensions["D"].width = 64
        sheet.column_dimensions["E"].width = 64

        # 添加数据
        for i, item in enumerate(items):
            TableManager.set_cell_value(sheet, i + 2, 1, item.get("file_path"))
            TableManager.set_cell_value(sheet, i + 2, 2, "\n".join(item.get("group")))
            TableManager.set_cell_value(sheet, i + 2, 3, item.get("src"))
            TableManager.set_cell_value(sheet, i + 2, 4, item.get("dst"))
            TableManager.set_cell_value(sheet, i + 2, 5, item.get("dst"))

        # 保存工作簿
        abs_path = f"{config.output_folder}/{Localizer.get().path_result_batch_correction}"
        os.makedirs(os.path.dirname(abs_path), exist_ok = True)
        book.save(abs_path)

        # 提示
        self.emit(Base.Event.TOAST, {
            "type": Base.ToastType.SUCCESS,
            "message": Localizer.get().task_success,
        })

    # 第二步点击事件
    def step_02_clicked(self, window: FluentWindow) -> None:
        config = Config().load()

        data_dict: dict[str, list[dict[str, str]]] = {}
        abs_path = f"{config.output_folder}/{Localizer.get().path_result_batch_correction}"
        try:
            # 数据处理
            book: openpyxl.Workbook = openpyxl.load_workbook(abs_path)
            sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

            # 跳过空表格
            if sheet.max_row == 0 or sheet.max_column == 0:
                raise Exception()

            for row in range(2, sheet.max_row + 1):
                file_path: str = sheet.cell(row = row, column = 1).value
                src: str = sheet.cell(row = row, column = 3).value
                dst: str = sheet.cell(row = row, column = 4).value
                fix: str = sheet.cell(row = row, column = 5).value

                # 跳过读取失败的行
                # 数据不存在时为 None，存在时可能是 str int float 等多种类型
                if file_path is None:
                    continue

                src = str(src) if src is not None else ""
                dst = str(dst) if dst is not None else ""
                fix = str(fix) if fix is not None else ""

                # 跳过无修改的行
                if fix == "" or fix == dst:
                    continue

                data_dict.setdefault(file_path, []).append({
                    "src": src,
                    "dst": dst,
                    "fix": fix,
                })
        except Exception:
            pass

        # 读取输入文件
        _, items = FileManager(config).read_from_path()
        items = [
            v for v in items
            if BatchCorrectionPage.FILE_NAME_WHITELIST.search(v.get_file_path()) is None
        ]

        # 有效性检查
        if len(data_dict) == 0 or len(items) == 0:
            self.emit(Base.Event.TOAST, {
                "type": Base.ToastType.ERROR,
                "message": Localizer.get().alert_no_data,
            })
            return None

        # 修正数据
        for item in items:
            src = item.get_src().replace("\r", "").replace("_x000D_", "")
            dst = item.get_dst().replace("\r", "").replace("_x000D_", "")
            file_path = item.get_file_path()
            if file_path in data_dict:
                for data in data_dict.get(file_path):
                    flag: bool = False
                    if item.get_file_type() in BatchCorrectionPage.SINGLE:
                        flag = dst == data.get("dst").replace("\r", "").replace("_x000D_", "")
                    elif item.get_file_type() in BatchCorrectionPage.DOUBLE:
                        flag = (
                            src == data.get("src").replace("\r", "").replace("_x000D_", "")
                            and dst == data.get("dst").replace("\r", "").replace("_x000D_", "")
                        )
                    if flag == True:
                        item.set_dst(self.auto_convert_line_break(item.get_src(), data.get("fix")))
                        break

        # 写入文件
        threading.Thread(
            target = self.write_to_path_task,
            args = (config, items),
        ).start()

    # 写入文件
    def write_to_path_task(self, config: Config, items: list[Item]) -> None:
        FileManager(config).write_to_path(items)
        shutil.rmtree(f"{config.output_folder}/{Localizer.get().path_bilingual}", ignore_errors = True)

        # 提示
        self.emit(Base.Event.TOAST, {
            "type": Base.ToastType.SUCCESS,
            "message": Localizer.get().task_success,
        })


    # 根据原文换行符对修正文本中的换行符进行转换
    def auto_convert_line_break(self, src: str, fix: str) -> str:
        if "_x000D_" not in src and "\r" not in src:
            return fix
        else:
            return fix.replace("\n", "_x000D_\n").replace("_x000D_\n", "\r\n")