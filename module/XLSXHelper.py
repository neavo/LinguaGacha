import openpyxl
import openpyxl.styles
import openpyxl.worksheet.worksheet
from openpyxl.worksheet.worksheet import Worksheet

class XLSXHelper():

    def __init__(self) -> None:
        super().__init__()

    # 设置单元格值
    @classmethod
    def set_cell_value(cls, sheet: Worksheet, row: int, column: int, value: str, font_size: int = 9) -> None:
        # 如果单元格内容以单引号 ' 开头，Excel 会将其视为普通文本而不是公式
        sheet.cell(row = row, column = column).value = value if value.startswith("=") == False else f"'{value}"
        sheet.cell(row = row, column = column).font = openpyxl.styles.Font(size = font_size)
        sheet.cell(row = row, column = column).alignment  = openpyxl.styles.Alignment(wrap_text = True, vertical = "center", horizontal = "left")