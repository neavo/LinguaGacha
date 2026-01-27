import json
import re
from enum import StrEnum
from functools import partial
from typing import Any
from typing import Iterable

import openpyxl
import openpyxl.styles
import openpyxl.worksheet.worksheet
from PyQt5.QtCore import QModelIndex
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFontMetrics
from PyQt5.QtWidgets import QTableWidgetItem
from qfluentwidgets import TableWidget

from widget.RuleWidget import RuleWidget


class TableManager:
    class Type(StrEnum):
        GLOSSARY = "GLOSSARY"
        REPLACEMENT = "REPLACEMENT"
        TEXT_PRESERVE = "TEXT_PRESERVE"

    ROW_NUMBER_WIDTHS = {
        Type.GLOSSARY: 44,
        Type.REPLACEMENT: 40,
        Type.TEXT_PRESERVE: 40,
    }

    def __init__(
        self, type: str, data: list[dict[str, str]], table: TableWidget
    ) -> None:
        super().__init__()

        # 初始化
        self.type = type
        self.data = data
        self.table = table

        # 更新中标识
        self.updating: bool = False

    # 重置
    def reset(self) -> None:
        self.data = []
        self.table.clearContents()
        self.table.horizontalHeader().setSortIndicator(-1, Qt.SortOrder.AscendingOrder)

    # 同步
    def sync(self) -> None:
        # 更新开始
        self.set_updating(True)

        # 去重
        duplicate_indices: set[int] = set()
        for i in range(len(self.data)):
            for k in range(i + 1, len(self.data)):
                x = self.data[i]
                y = self.data[k]
                if x.get("src") == y.get("src"):
                    if x.get("dst") != "" and y.get("dst") == "":
                        duplicate_indices.add(k)
                    elif (
                        x.get("dst") == ""
                        and y.get("dst") == ""
                        and x.get("info") != ""
                        and y.get("info") == ""
                    ):
                        duplicate_indices.add(k)
                    elif (
                        x.get("dst") == ""
                        and y.get("dst") == ""
                        and x.get("regex") != ""
                        and y.get("regex") == ""
                    ):
                        duplicate_indices.add(k)
                    else:
                        duplicate_indices.add(i)
        self.data = [v for i, v in enumerate(self.data) if i not in duplicate_indices]

        # 填充表格
        self.table.setRowCount(max(20, len(self.data) + 8))
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item is not None:
                    item.setText("")
                else:
                    self.table.setItem(row, col, self.generate_item(col))

        # 遍历表格
        if self.type == __class__.Type.GLOSSARY:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("dst", ""))
                    elif col == 2:
                        self.table.item(row, col).setText(v.get("info", ""))
                    elif col == 3:
                        rule_widget = RuleWidget(
                            show_regex=False,
                            show_case_sensitive=True,
                            case_sensitive_enabled=v.get("case_sensitive", False),
                            on_changed=partial(self.on_rule_changed, row, v),
                        )
                        self.table.setCellWidget(row, col, rule_widget)
        elif self.type == __class__.Type.REPLACEMENT:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("dst", ""))
                    elif col == 2:
                        rule_widget = RuleWidget(
                            show_regex=True,
                            show_case_sensitive=True,
                            regex_enabled=v.get("regex", False),
                            case_sensitive_enabled=v.get("case_sensitive", False),
                            on_changed=partial(self.on_rule_changed, row, v),
                        )
                        self.table.setCellWidget(row, col, rule_widget)
        elif self.type == __class__.Type.TEXT_PRESERVE:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("info", ""))

        self.update_vertical_headers()

        # 更新结束
        self.set_updating(False)

    def update_vertical_headers(self) -> None:
        min_width = self.ROW_NUMBER_WIDTHS.get(self.type, 40)
        row_count = max(self.table.rowCount(), 1)
        digits = len(str(row_count))
        metrics = QFontMetrics(self.table.verticalHeader().font())
        text_width = metrics.horizontalAdvance("9" * digits)
        self.table.verticalHeader().setFixedWidth(max(min_width, text_width + 16))
        for row in range(self.table.rowCount()):
            item = self.table.verticalHeaderItem(row)
            label = str(row + 1)
            if item is None:
                item = QTableWidgetItem(label)
                self.table.setVerticalHeaderItem(row, item)
            else:
                item.setText(label)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    # 导出
    def export(self, path: str) -> None:
        # 新建工作表
        book: openpyxl.Workbook = openpyxl.Workbook()
        sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

        # 设置表头
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 24
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 24
        TableManager.set_cell_value(sheet, 1, 1, "src", 10)
        TableManager.set_cell_value(sheet, 1, 2, "dst", 10)
        TableManager.set_cell_value(sheet, 1, 3, "info", 10)
        TableManager.set_cell_value(sheet, 1, 4, "regex", 10)
        TableManager.set_cell_value(sheet, 1, 5, "case_sensitive", 10)

        # 将数据写入工作表
        for row, item in enumerate(self.data):
            TableManager.set_cell_value(sheet, row + 2, 1, item.get("src", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 2, item.get("dst", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 3, item.get("info", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 4, item.get("regex", ""), 10)
            TableManager.set_cell_value(
                sheet, row + 2, 5, item.get("case_sensitive", ""), 10
            )

        # 保存工作簿
        book.save(f"{path}.xlsx")

        # 保存为 JSON
        with open(f"{path}.json", "w", encoding="utf-8") as writer:
            writer.write(json.dumps(self.data, indent=4, ensure_ascii=False))

    # 搜索
    def search(self, keyword: str, start: int) -> int:
        result: int = -1
        keyword = keyword.lower()

        # 从指定位置开始搜索
        for i, entry in enumerate(self.data):
            if i <= start:
                continue
            if any(keyword in v.lower() for v in entry.values() if isinstance(v, str)):
                result = i
                break

        # 如果未找到则从头开始搜索
        if result == -1:
            for i, entry in enumerate(self.data):
                if i >= start:
                    continue
                if any(
                    keyword in v.lower() for v in entry.values() if isinstance(v, str)
                ):
                    result = i
                    break

        return result

    @staticmethod
    def build_table_matches(
        table: TableWidget,
        keyword: str,
        use_regex: bool,
        columns: Iterable[int],
    ) -> tuple[list[int], set[int]]:
        matches: list[int] = []
        empty_rows: set[int] = set()

        if use_regex:
            try:
                pattern = re.compile(keyword, re.IGNORECASE)
            except re.error:
                return [], set()
            keyword_lower = ""
        else:
            pattern = None
            keyword_lower = keyword.lower()

        for row in range(table.rowCount()):
            texts = []
            for col in columns:
                item = table.item(row, col)
                if not item:
                    continue
                text = item.text().strip()
                if text:
                    texts.append(text)

            if not texts:
                empty_rows.add(row)
                continue

            if not keyword:
                continue

            if pattern:
                if any(pattern.search(text) for text in texts):
                    matches.append(row)
            else:
                if any(keyword_lower in text.lower() for text in texts):
                    matches.append(row)

        return matches, empty_rows

    @staticmethod
    def find_current_match_index(matches: list[int], row: int) -> int:
        if row in matches:
            return matches.index(row)
        return -1

    @staticmethod
    def pick_next_match(matches: list[int], current_row: int, reverse: bool) -> int:
        if not matches:
            return -1

        if reverse:
            prev_matches = [m for m in matches if m < current_row]
            if prev_matches:
                return prev_matches[-1]
            return matches[-1]

        next_matches = [m for m in matches if m > current_row]
        if next_matches:
            return next_matches[0]
        return matches[0]

    # 获取数据
    def get_data(self) -> list[dict[str, str]]:
        return self.data

    # 设置数据
    def set_data(self, data: list[dict[str, str]]) -> None:
        self.data = data

    # 获取更新中标识
    def get_updating(self) -> bool:
        return self.updating

    # 设置更新中标识
    def set_updating(self, updating: bool) -> None:
        self.updating = updating

    # 规则变更回调
    def on_rule_changed(
        self,
        row: int,
        data_ref: dict[str, str | bool],
        regex: bool,
        case_sensitive: bool,
    ) -> None:
        if self.type == __class__.Type.REPLACEMENT:
            data_ref["regex"] = regex

        data_ref["case_sensitive"] = case_sensitive

        # 触发信号
        self.table.itemChanged.emit(self.table.item(row, 0))

    # 生成新的条目
    def generate_item(self, col: int) -> QTableWidgetItem:
        item = QTableWidgetItem("")
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        if self.type == __class__.Type.GLOSSARY:
            # 规则列不可编辑
            if col == 3:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        elif self.type == __class__.Type.REPLACEMENT:
            # 规则列不可编辑
            if col == 2:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        elif self.type == __class__.Type.TEXT_PRESERVE:
            pass

        return item

    # 删除行事件
    def delete_row(self) -> None:
        selected_index = self.table.selectedIndexes()

        # 有效性检验
        if selected_index is None or len(selected_index) == 0:
            return

        # 逆序删除并去重以避免索引错误
        for row in sorted({item.row() for item in selected_index}, reverse=True):
            self.table.removeRow(row)

        # 删除行后同步数据
        self.set_data([])
        self.append_data_from_table()

    # 切换正则事件
    def switch_regex(self) -> None:
        selected_index: list[QModelIndex] = self.table.selectedIndexes()

        # 有效性检验
        if selected_index is None or len(selected_index) == 0:
            return

        # 切换正则模式
        for row in {index.row() for index in selected_index}:
            item = self.table.item(row, 2)
            if item is None:
                item = QTableWidgetItem()
                self.table.setItem(row, 2, item)
            if item.text().strip() != "✅":
                item.setText("✅")
            else:
                item.setText("")

    # 获取行数据
    def get_entry_by_row(self, row: int) -> dict[str, str | bool]:
        items: list[QTableWidgetItem] = [
            self.table.item(row, col) for col in range(self.table.columnCount())
        ]

        if self.type == __class__.Type.GLOSSARY:
            # 从规则列的 RuleWidget 获取 case_sensitive 状态
            rule_widget = self.table.cellWidget(row, 3)
            case_sensitive = (
                rule_widget.get_case_sensitive_enabled()
                if isinstance(rule_widget, RuleWidget)
                else False
            )

            return {
                "src": items[0].text().strip()
                if isinstance(items[0], QTableWidgetItem)
                else "",
                "dst": items[1].text().strip()
                if isinstance(items[1], QTableWidgetItem)
                else "",
                "info": items[2].text().strip()
                if isinstance(items[2], QTableWidgetItem)
                else "",
                "case_sensitive": case_sensitive,
            }
        elif self.type == __class__.Type.REPLACEMENT:
            # 从规则列的 RuleWidget 获取 regex 和 case_sensitive 状态
            rule_widget = self.table.cellWidget(row, 2)
            regex = (
                rule_widget.get_regex_enabled()
                if isinstance(rule_widget, RuleWidget)
                else False
            )
            case_sensitive = (
                rule_widget.get_case_sensitive_enabled()
                if isinstance(rule_widget, RuleWidget)
                else False
            )

            return {
                "src": items[0].text().strip()
                if isinstance(items[0], QTableWidgetItem)
                else "",
                "dst": items[1].text().strip()
                if isinstance(items[1], QTableWidgetItem)
                else "",
                "info": "",
                "regex": regex,
                "case_sensitive": case_sensitive,
            }
        elif self.type == __class__.Type.TEXT_PRESERVE:
            return {
                "src": items[0].text().strip()
                if isinstance(items[0], QTableWidgetItem)
                else "",
                "info": items[1].text().strip()
                if isinstance(items[1], QTableWidgetItem)
                else "",
            }

    # 从表格加载数据
    def append_data_from_table(self) -> None:
        for row in range(self.table.rowCount()):
            entry: dict[str, str | bool] = self.get_entry_by_row(row)
            if entry.get("src") != "":
                self.data.append(entry)

    # 从文件加载数据
    def append_data_from_file(self, path: str) -> None:
        result: list[dict[str, str]] = []

        if path.lower().endswith(".json"):
            result = self.load_from_json_file(path)
        elif path.lower().endswith(".xlsx"):
            result = self.load_from_xlsx_file(path)

        # 合并数据并去重
        self.data.extend(result)
        self.data = list({v["src"]: v for v in self.data}.values())

    # 从 json 文件加载数据
    def load_from_json_file(self, path: str) -> list[dict[str, str]]:
        result: list[dict[str, str]] = []

        # 读取文件
        inputs = []
        with open(path, "r", encoding="utf-8-sig") as reader:
            inputs: dict[str, str] | list[dict[str, str]] = json.load(reader)

        # 标准字典列表
        # [
        #     {
        #         "key": "value",
        #         "key": "value",
        #         "key": "value",
        #     }
        # ]
        if isinstance(inputs, list):
            for entry in inputs:
                # 格式校验
                if not isinstance(entry, dict):
                    continue
                if "src" not in entry:
                    continue

                src: str = entry.get("src", "").strip()
                if src != "":
                    result.append(
                        {
                            "src": src,
                            "dst": entry.get("dst", "").strip(),
                            "info": entry.get("info", "").strip(),
                            "regex": entry.get("regex", False),
                            "case_sensitive": entry.get("case_sensitive", False),
                        }
                    )

        # Actors.json
        # [
        #     null,
        #     {
        #         "id": 1,
        #         "name": "レナリス",
        #         "nickname": "ローズ娼館の娼婦",
        #     },
        # ]
        if isinstance(inputs, list):
            for entry in inputs:
                # 格式校验
                if not isinstance(entry, dict):
                    continue
                if not isinstance(entry.get("id"), int):
                    continue

                id: int = entry.get("id", -1)
                name: str = entry.get("name", "").strip()
                nickname: str = entry.get("nickname", "").strip()

                # 添加数据
                if name != "":
                    result.append(
                        {
                            "src": f"\\n[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )
                    result.append(
                        {
                            "src": f"\\N[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )
                if nickname != "":
                    result.append(
                        {
                            "src": f"\\nn[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )
                    result.append(
                        {
                            "src": f"\\NN[{id}]",
                            "dst": name,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )

        # 标准 KV 字典
        # {
        #     "ダリヤ": "达莉雅"
        # }
        if isinstance(inputs, dict):
            for k, v in inputs.items():
                # 格式校验
                if not isinstance(k, str):
                    continue

                src: str = k.strip()
                dst: str = str(v).strip() if v is not None else ""
                if src != "":
                    result.append(
                        {
                            "src": src,
                            "dst": dst,
                            "info": "",
                            "regex": False,
                            "case_sensitive": False,
                        }
                    )

        return result

    # 从 xlsx 文件加载数据
    def load_from_xlsx_file(self, path: str) -> list[dict]:
        result: list[dict[str, str]] = []

        sheet = openpyxl.load_workbook(path).active
        for row in range(1, sheet.max_row + 1):
            # 读取每一行的数据
            data: list[str] = [
                __class__.get_cell_value(sheet, row, col) for col in range(1, 6)
            ]

            # 格式校验
            if len(data) == 0 or data[0] is None:
                continue

            src: str = data[0]
            dst: str = data[1]
            info: str = data[2]
            regex: bool = data[3].lower() == "true" if len(data) > 3 else False
            case_sensitive: bool = data[4].lower() == "true" if len(data) > 4 else False

            if src == "src" and dst == "dst":
                continue

            # 添加数据
            if src != "":
                result.append(
                    {
                        "src": src,
                        "dst": dst,
                        "info": info,
                        "regex": regex,
                        "case_sensitive": case_sensitive,
                    }
                )

        return result

    # 设置单元格值
    @classmethod
    def get_cell_value(
        cls, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int
    ) -> str:
        value = sheet.cell(row=row, column=column).value

        # 强制转换为字符串
        if value is None:
            result = ""
        else:
            result = str(value)

        return result.strip()

    # 设置单元格值
    @classmethod
    def set_cell_value(
        cls,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        column: int,
        value: Any,
        font_size: int = 9,
    ) -> None:
        if value is None:
            value = ""
        # 如果单元格内容以单引号 ' 开头，Excel 会将其视为普通文本而不是公式
        elif isinstance(value, str) and value.startswith("="):
            value = "'" + value

        sheet.cell(row=row, column=column).value = value
        sheet.cell(row=row, column=column).font = openpyxl.styles.Font(size=font_size)
        sheet.cell(row=row, column=column).alignment = openpyxl.styles.Alignment(
            wrap_text=True, vertical="center", horizontal="left"
        )
