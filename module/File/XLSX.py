import os

import openpyxl
import openpyxl.worksheet.worksheet

from base.Base import Base
from module.Cache.CacheItem import CacheItem
from module.XLSXHelper import XLSXHelper

class XLSX(Base):

    def __init__(self, config: dict) -> None:
        super().__init__()

        # 初始化
        self.config: dict = config
        self.input_path: str = config.get("input_folder")
        self.output_path: str = config.get("output_folder")
        self.source_language: str = config.get("source_language")
        self.target_language: str = config.get("target_language")

    # 读取
    def read_from_path(self, abs_paths: list[str]) -> list[CacheItem]:
        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, self.input_path)

            # 数据处理
            book: openpyxl.Workbook = openpyxl.load_workbook(abs_path)
            sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

            # 跳过空表格
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue

            # 判断是否为 WOLF 翻译表格文件
            if self.is_wold_xlsx(sheet):
                continue

            for row in range(1, sheet.max_row + 1):
                src = sheet.cell(row = row, column = 1).value
                dst = sheet.cell(row = row, column = 2).value

                # 跳过读取失败的行
                # 数据不存在时为 None，存在时可能是 str int float 等多种类型
                if src is None:
                    continue

                src = str(src)
                dst = str(dst) if dst is not None else ""
                if src == "":
                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.TranslationStatus.EXCLUDED,
                        })
                    )
                elif dst != "" and src != dst:
                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.TranslationStatus.TRANSLATED_IN_PAST,
                        })
                    )
                else:
                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.TranslationStatus.UNTRANSLATED,
                        })
                    )

        return items

    # 写入
    def write_to_path(self, items: list[CacheItem]) -> None:
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.XLSX
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            # 按行号排序
            items = sorted(items, key = lambda x: x.get_row())

            # 新建工作表
            book: openpyxl.Workbook = openpyxl.Workbook()
            sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

            # 设置表头
            sheet.column_dimensions["A"].width = 64
            sheet.column_dimensions["B"].width = 64

            # 将数据写入工作表
            for item in items:
                row: int = item.get_row()
                XLSXHelper.set_cell_value(sheet, row, column = 1, value = item.get_src())
                XLSXHelper.set_cell_value(sheet, row, column = 2, value = item.get_dst())

            # 保存工作簿
            abs_path = f"{self.output_path}/{rel_path}"
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            book.save(abs_path)

    # 是否为 WOLF 翻译表格文件
    def is_wold_xlsx(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        value: str = sheet.cell(row = 1, column = 1).value
        if not isinstance(value, str) or "code" not in value.lower():
            return False

        value: str = sheet.cell(row = 1, column = 2).value
        if not isinstance(value, str) or "flag" not in value.lower():
            return False

        value: str = sheet.cell(row = 1, column = 3).value
        if not isinstance(value, str) or "type" not in value.lower():
            return False

        value: str = sheet.cell(row = 1, column = 4).value
        if not isinstance(value, str) or "info" not in value.lower():
            return False

        return True