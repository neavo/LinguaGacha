import os

import openpyxl
import openpyxl.styles
from openpyxl.worksheet.worksheet import Worksheet

from base.Base import Base
from base.BaseLanguage import BaseLanguage
from model.Item import Item
from module.Config import Config
from module.Storage.AssetStore import AssetStore
from module.Storage.PathStore import PathStore
from module.Storage.StorageContext import StorageContext
from module.TableManager import TableManager


class WOLFXLSX(Base):
    BLACKLIST_EXT: tuple[str, ...] = (
        ".mp3",
        ".wav",
        ".ogg",
        "mid",
        ".png",
        ".jpg",
        ".jpeg",
        ".gif",
        ".psd",
        ".webp",
        ".heif",
        ".heic",
        ".avi",
        ".mp4",
        ".webm",
        ".txt",
        ".7z",
        ".gz",
        ".rar",
        ".zip",
        ".json",
        ".sav",
        ".mps",
        ".ttf",
        ".otf",
        ".woff",
    )

    FILL_COLOR_WHITELIST: tuple = (
        9,  # 白色
    )

    FILL_COLOR_BLACKLIST: tuple = (
        44,  # 蓝色
        47,  # 土黄
        55,  # 灰色
    )

    COL_SRC_TEXT = 6
    COL_DST_TEXT = 7

    def __init__(self, config: Config) -> None:
        super().__init__()

        # 初始化
        self.config = config
        self.source_language: BaseLanguage.Enum = config.source_language
        self.target_language: BaseLanguage.Enum = config.target_language

    # 读取
    def read_from_path(self, abs_paths: list[str], input_path: str) -> list[Item]:
        items: list[Item] = []
        for abs_path in abs_paths:
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            book: openpyxl.Workbook = openpyxl.load_workbook(abs_path)
            sheet = book.active

            # Ensure it is a Worksheet
            if not isinstance(sheet, Worksheet):
                continue

            # 跳过空表格
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue

            # 判断是否为 WOLF 翻译表格文件
            if not self.is_wolf_xlsx(sheet):
                continue

            for row in range(2, sheet.max_row + 1):
                src_val = sheet.cell(row=row, column=self.COL_SRC_TEXT).value

                # 跳过读取失败的行
                # 数据不存在时为 None，存在时可能是 str int float 等多种类型
                if src_val is None:
                    continue

                src: str = str(src_val)
                dst_val = sheet.cell(row=row, column=self.COL_DST_TEXT).value
                dst: str = str(dst_val) if dst_val is not None else ""

                status = Base.ProjectStatus.NONE

                if (
                    src == ""
                    or self.get_fg_color_index(sheet, row, self.COL_SRC_TEXT)
                    not in self.FILL_COLOR_WHITELIST
                ):
                    status = Base.ProjectStatus.EXCLUDED
                elif dst != "" and src != dst:
                    status = Base.ProjectStatus.PROCESSED_IN_PAST

                items.append(
                    Item.from_dict(
                        {
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": Item.FileType.WOLFXLSX,
                            "file_path": rel_path,
                            "text_type": Item.TextType.WOLF,
                            "status": status,
                        }
                    )
                )

        return items

    # 写入
    def write_to_path(self, items: list[Item]) -> None:
        # 获取输出目录
        output_path = PathStore.get_translated_path()
        db = StorageContext.get().get_db()

        target = [
            item for item in items if item.get_file_type() == Item.FileType.WOLFXLSX
        ]

        # 按文件路径分组
        group: dict[str, list[Item]] = {}
        for item in target:
            group.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in group.items():
            # 按行号排序
            items = sorted(items, key=lambda x: x.get_row())

            # 目标文件绝对路径
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)

            # 尝试从数据库恢复原始文件
            restored = False
            if db:
                asset_data = db.get_asset(rel_path)
                if asset_data:
                    AssetStore.decompress_to_file(asset_data, abs_path)
                    restored = True

            if restored:
                # 加载恢复的文件
                book: openpyxl.Workbook = openpyxl.load_workbook(abs_path)
                sheet = book.active
            else:
                # 回退方案：新建工作表（仅当无法恢复时）
                book: openpyxl.Workbook = openpyxl.Workbook()
                sheet = book.active
                # 设置表头
                if isinstance(sheet, Worksheet):
                    sheet.column_dimensions["A"].width = 64
                    sheet.column_dimensions["B"].width = 64

            if not isinstance(sheet, Worksheet):
                continue

            # 将数据写入工作表
            for item in items:
                row: int = item.get_row()
                TableManager.set_cell_value(
                    sheet, row, column=self.COL_SRC_TEXT, value=item.get_src()
                )
                TableManager.set_cell_value(
                    sheet, row, column=self.COL_DST_TEXT, value=item.get_dst()
                )

            # 保存工作簿
            book.save(abs_path)

    # 是否为 WOLF 翻译表格文件
    def is_wolf_xlsx(self, sheet: Worksheet) -> bool:
        headers = {1: "code", 2: "flag", 3: "type", 4: "info"}
        for col, expected in headers.items():
            value = sheet.cell(row=1, column=col).value
            if not isinstance(value, str) or expected not in value.lower():
                return False
        return True

    # 获取单元格填充颜色索引
    def get_fg_color_index(self, sheet: Worksheet, row: int, column: int) -> int:
        fill = sheet.cell(row=row, column=column).fill
        if fill.fill_type is None:
            return -1

        fg_color = fill.fgColor
        if (
            fg_color
            and isinstance(fg_color, openpyxl.styles.Color)
            and fg_color.type == "indexed"
        ):
            return fg_color.indexed

        return -1
